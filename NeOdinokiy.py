from aiogram import Bot, Dispatcher, executor, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import Message
from openpyxl import load_workbook
import logging
import sqlite3

path = 'data.xlsx'
wb = load_workbook(filename=path)
ws = wb.get_sheet_by_name('data')

API_TOKEN = '6162485968:AAEMRUmRJ-uyReEgyaEjvoi0K9Fa4uE1XpQ'
ADMIN1 = 337667307
ADMIN2 = 413981336


kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
kb.add(types.InlineKeyboardButton(text="Рассылка"))
kb.add(types.InlineKeyboardButton(text="Статистика"))
kb.add(types.InlineKeyboardButton(text="Написать"))

logging.basicConfig(level=logging.INFO)
storage = MemoryStorage()
bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot, storage=storage)

conn = sqlite3.connect('database.db')
cur = conn.cursor()
cur.execute("""CREATE TABLE IF NOT EXISTS users(user_id INTEGER, block INTEGER);""")
conn.commit()


class dialog(StatesGroup):
    result_spam = State()
    chat = State()
    spam = State()
    blacklist = State()
    whitelist = State()


@dp.message_handler(commands=['start'])
async def start(message: Message):
    cur = conn.cursor()
    cur.execute(f"SELECT block FROM users WHERE user_id = {message.chat.id}")
    result = cur.fetchone()
    if message.from_user.id == ADMIN1 or message.from_user.id == ADMIN2:
        await message.answer('Добро пожаловать в Админ-Панель! Выберите действие на клавиатуре', reply_markup=kb)
    else:
        if result is None:
            cur = conn.cursor()
            cur.execute(f'''SELECT * FROM users WHERE (user_id="{message.from_user.id}")''')
            entry = cur.fetchone()
            if entry is None:
                cur.execute(f'''INSERT INTO users VALUES ('{message.from_user.id}', '0')''')
                conn.commit()
                chat_id = message.from_user.id
                await message.answer(f"Привет, к сожаление подача заявок окончена(")
                await message.answer_sticker('CAACAgEAAxkBAAEH2-1j9j6BHR3mjoq1WJjhcynAxTkYtgACIAADLEj-P8CLoqSC0wABZi4E')
            else:
                await message.answer("Да-да, я уже тут)")






@dp.message_handler(state=dialog.result_spam)
async def result_s(message: Message, state: FSMContext):
    if message.from_user.id == ADMIN1 or message.from_user.id == ADMIN2:
        if message.text == 'Назад':
            await message.answer('Главное меню', reply_markup=kb)
            await state.finish()
        else:
            for x in range(137, 2, -1):
                name = ws.cell(row=x, column=2).value #
                link = ws.cell(row=x, column=3).value
                faculty = ws.cell(row=x, column=5).value #
                live = ws.cell(row=x, column=9).value
                chrono = ws.cell(row=x, column=10).value #
                high = ws.cell(row=x, column=12).value
                zodiac = ws.cell(row=x, column=15).value
                music = ws.cell(row=x, column=16).value
                vert = ws.cell(row=x, column=17).value
                hobby = ws.cell(row=x, column=18).value
                best = ws.cell(row=x, column=19).value #
                gippocrat = ws.cell(row=x, column=23).value
                book = ws.cell(row=x, column=26).value
                fav_music = ws.cell(row=x, column=27).value
                word_about = ws.cell(row=x, column=28).value
                clean_or_not = ws.cell(row=x, column=29).value
                photo = ws.cell(row=x, column=30).value
                message_to_partner = ws.cell(row=x, column=31).value
                partner = ws.cell(row=x, column=32).value
                if await bot.send_message(int(partner), 'Приветик, долго не виделись, вот партнер которого мы тебе подобрали:'):
                    await bot.send_sticker(int(partner), 'CAACAgEAAxkBAAEHmoJj4HvCdX0keydJszM9hzK7ahdzVAACLgADLEj-P9T0k8ntWD-MLgQ')
                    await bot.send_message(int(partner), photo + '\nИмя:' + name + '\nФакультет:' + faculty + '\nРост:' + high + '\nЛучшие черты:' + best)
                    await bot.send_message(int(partner), 'Кем является:' + vert + '\nПо гиппократу:' + gippocrat + '\nХронотип:' + chrono + '\nЗнак зодиака:' + zodiac + '\nВыбирает' + clean_or_not)
                    await bot.send_message(int(partner), 'Проживает в:' + live +  '\nХобби:' + hobby + '\nЛюбимые книги:' + book +'\nЛюбимые жанры музыки:' + music + '\nЛюбимый трек' + fav_music)
                    await bot.send_message(int(partner), 'Одним словом):' + word_about)
                    await bot.send_message(int(partner), 'Доверься нашим специалистам в области соулмейтов и не только) не бойся и не стесняйся)\nИди пиши скорей\n' + link)
                    if message_to_partner:
                        await bot.send_message(int(partner), 'А, и послание специально для тебя')
                        await bot.send_message(int(partner), message_to_partner)
                else:
                    await bot.send_message(ADMIN1, link)
            await message.answer('Рассылка завершена', reply_markup=kb)
            await state.finish()


@dp.message_handler(content_types=['text'], text='Рассылка')
async def spam(message: Message):
    if message.from_user.id == ADMIN1 or message.from_user.id == ADMIN2:
        await dialog.spam.set()
        await message.answer('Напиши текст рассылки')


@dp.message_handler(state=dialog.spam)
async def start_spam(message: Message, state: FSMContext):
    if message.from_user.id == ADMIN1 or message.from_user.id == ADMIN2:
        if message.text == 'Назад':
            await message.answer('Главное меню', reply_markup=kb)
            await state.finish()
        else:
            cur = conn.cursor()
            cur.execute(f'''SELECT user_id FROM users''')
            spam_base = cur.fetchall()
            for z in range(len(spam_base) - 1, 1,-1):
                await bot.send_message(spam_base[z][0], message.text)
            await message.answer('Рассылка завершена', reply_markup=kb)
            await state.finish()



@dp.message_handler(content_types=['text'], text='Статистика')
async def hfandler(message: types.Message, state: FSMContext):
    if message.from_user.id == ADMIN1 or message.from_user.id == ADMIN2:
        cur = conn.cursor()
        cur.execute('''select * from users''')
        results = cur.fetchall()
        await message.answer(f'Людей которые когда либо заходили в бота: {len(results)}')



wb.close()
if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)

